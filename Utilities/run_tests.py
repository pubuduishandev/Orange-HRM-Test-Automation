import sys
import os
import subprocess
import json
import datetime
from rich.console import Console
from rich.table import Table
from shutil import get_terminal_size
from rich.text import Text
import data_loader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Step 1 - CMD Customization
# Disable the maximize button
if sys.platform == "win32":
    try:
        from Utilities import disable_maximize
    except ImportError:
        import disable_maximize

console = Console()
start_time = datetime.datetime.now()

# Change the CMD title
if sys.platform == "win32":
    os.system("title ORANGE HRM TEST AUTOMATION V1.0.0.1")

# Step 2: Display start info
# Get terminal width
terminal_width = get_terminal_size().columns

# Styled multi-line header using gradient and padding
header_text = Text()
header_text.append("ORANGE HRM TEST AUTOMATION".center(terminal_width), style="bold white")
header_text.append("\n", style="bold")
header_text.append("Version OS 5.7".center(terminal_width), style="bold orange3")
header_text.append("\n" + "".center(terminal_width), style="bold")

header_text.append("".center(terminal_width) + "\n", style="bold")
header_text.append(f"Test Started At: {start_time.strftime('%Y-%m-%d %H:%M:%S')}".center(terminal_width), style="bold cyan")

console.print(header_text)

# Step 3: Run pytest and generate a JSON result file
timestamp_str = start_time.strftime("%Y-%m-%d_%H-%M-%S")
result_file = f"Reports/JSON/test_{timestamp_str}_log_report.json"
cmd = f"pytest testCases/ --json-report --json-report-file={result_file} --tb=no -q > nul"
subprocess.run(cmd, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

end_time = datetime.datetime.now()
duration = round((end_time - start_time).total_seconds(), 2)

# Handle report missing
if not os.path.exists(result_file):
    console.print("[red]Failed to generate test report.[/red]")
    exit(1)

with open(result_file) as f:
    data = json.load(f)

# Step 4: nodeid → loader mapping
def get_test_data_by_nodeid(nodeid: str):
    if "test_home.py::test_home_page_title" in nodeid:
        return data_loader.load_urls_from_json()
    elif "test_login.py::test_login_success" in nodeid:
        return data_loader.load_users_from_json()
    elif "test_leave.py::test_navigate_to_leave" in nodeid:
        return data_loader.load_side_nav_menu_items_from_json()
    elif "test_logout.py::test_logout_functionality" in nodeid:
        return data_loader.load_profile_menu_items_from_json()
    else:
        return []

# Step 5: Show test results
tests = data.get("tests", [])
summary = data.get("summary", {})

console.print("".center(terminal_width) + "\n", style="bold")
console.print("TEST SUMMARY".center(terminal_width), style="bold white")

# Cache test data sets per test nodeid
test_data_cache = {}
test_data_index = {}

table = Table(show_lines=True)
table.add_column("Test ID", style="cyan", justify="left")
table.add_column("Test Case", style="white", justify="left")
table.add_column("Test Data Index", style="white", justify="left")
table.add_column("Test Data", style="white", justify="left")
table.add_column("Duration", style="bold", justify="left")
table.add_column("Result", style="bold", justify="left")

for i, test in enumerate(tests, 1):
    test_id = str(i).zfill(2)
    full_nodeid = test.get("nodeid", "N/A")
    test_case = full_nodeid.split("[")[0]

    setup_duration = test.get("setup", {}).get("duration", 0.0)
    call_duration = test.get("call", {}).get("duration", 0.0)
    teardown_duration = test.get("teardown", {}).get("duration", 0.0)

    total_duration = setup_duration + call_duration + teardown_duration
    duration_str = f"{round(total_duration, 2)}s"

    outcome = test.get("outcome", "unknown").upper()
    result_color = "green" if outcome == "PASSED" else "red"

    # Cache logic (now works properly)
    if test_case not in test_data_cache:
        test_data_cache[test_case] = get_test_data_by_nodeid(test_case)
        test_data_index[test_case] = 0

    test_data_list = test_data_cache[test_case]
    index = test_data_index[test_case]

    sample_data = test_data_list[index] if index < len(test_data_list) else {}
    sample_data_str = ', '.join(f"{k}: {v}" for k, v in sample_data.items()) if isinstance(sample_data, dict) else str(sample_data)
    test_data_count = len(test_data_list)

    # Properly increment the index
    test_data_index[test_case] += 1

    table.add_row(
        test_id,
        full_nodeid,
        f"{str(test_data_index[test_case])} of {test_data_count}",
        sample_data_str,
        duration_str,
        f"[{result_color}]{outcome}[/{result_color}]"
    )
console.print(table)

# Step 6: Stats summary
# Prepare raw summary text (no Rich styles yet)
raw_summary = (
    f"● Passed: {summary.get('passed', 0)}   "
    f"● Failed: {summary.get('failed', 0)}   "
    f"● Skipped: {summary.get('skipped', 0)}   "
    f"● Total: {summary.get('total', 0)}"
)

# Center the raw text
centered_summary = raw_summary.center(terminal_width)

# Now create a Rich Text object and style each word in-place
styled_summary = Text(centered_summary)
styled_summary.stylize("green", centered_summary.find("● Passed"), centered_summary.find("● Passed") + len(f"● Passed: {summary.get('passed', 0)}"))
styled_summary.stylize("red", centered_summary.find("● Failed"), centered_summary.find("● Failed") + len(f"● Failed: {summary.get('failed', 0)}"))
styled_summary.stylize("yellow", centered_summary.find("● Skipped"), centered_summary.find("● Skipped") + len(f"● Skipped: {summary.get('skipped', 0)}"))
styled_summary.stylize("cyan", centered_summary.find("● Total"), centered_summary.find("● Total") + len(f"● Total: {summary.get('total', 0)}"))

# Print it centered and styled
console.print("RESULTS OVERVIEW".center(terminal_width), style="bold white")
console.print(styled_summary)

# Step 7: Test Automation Execution Summary
console.print("".center(terminal_width) + "\n", style="bold")
console.print("Test Automation Execution Summary".center(terminal_width), style="bold white")

duration_str = str(end_time - start_time)

execution_summary = (
    f"Start: {start_time.strftime('%Y-%m-%d %H:%M:%S')}   " 
    f"End: {end_time.strftime('%Y-%m-%d %H:%M:%S')}   " 
    f"Duration: {duration}s ({duration_str})"
)

centered_execution_summary = execution_summary.center(terminal_width)
console.print(centered_execution_summary, style="bold white")

# Step 8: Display end info
console.print("".center(terminal_width) + "\n", style="bold")
console.print("***TEST END***".center(terminal_width), style="bold cyan")

console.print("".center(terminal_width) + "\n", style="bold")
console.print("Developed By".center(terminal_width), style="bold white")
console.print("© 2025 Pubudu Ishan Wickrama Arachchi | All Rights Reserved.".center(terminal_width), style="bold white")


# Step 9: Export test data to Excel (after Step 5)
excel_dir = os.path.join("Reports", "Excel")
os.makedirs(excel_dir, exist_ok=True)
excel_file_path = os.path.join(excel_dir, f"test_{timestamp_str}_log_report.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Test Summary"

headers = ["Test ID", "Test Case", "Test Data Index", "Test Data", "Duration", "Result"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment_center = Alignment(horizontal="center", vertical="center")

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    ws.column_dimensions[get_column_letter(col_num)].width = 30

# Reset test_data_index for export consistency
test_data_index_export = {k: 0 for k in test_data_cache.keys()}

for i, test in enumerate(tests, 1):
    full_nodeid = test.get("nodeid", "N/A")
    test_case = full_nodeid.split("[")[0]

    setup_duration = test.get("setup", {}).get("duration", 0.0)
    call_duration = test.get("call", {}).get("duration", 0.0)
    teardown_duration = test.get("teardown", {}).get("duration", 0.0)
    total_duration = setup_duration + call_duration + teardown_duration
    duration_str = f"{round(total_duration, 2)}s"

    outcome = test.get("outcome", "unknown").upper()

    # Use cached data and index from console step
    test_data_list = test_data_cache.get(test_case, [])
    index = test_data_index_export.get(test_case, 0)
    test_data_count = len(test_data_list)

    sample_data = test_data_list[index] if index < test_data_count else {}
    sample_data_str = ', '.join(f"{k}: {v}" for k, v in sample_data.items()) if isinstance(sample_data, dict) else str(sample_data)

    test_data_index_export[test_case] += 1

    row = [
        str(i).zfill(2),
        full_nodeid,
        f"{test_data_index_export[test_case]} of {test_data_count}",
        sample_data_str,
        duration_str,
        outcome,
    ]
    ws.append(row)

wb.save(excel_file_path)
